#!/usr/bin/env python
"""
Script para generar Excel de prueba para el Módulo 4 de Conciliación
Ejecutar: python generar_excel_prueba.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def crear_excel_prueba():
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación Noviembre 2025"

    # Configurar encabezados
    headers = ["Legajo", "Nombre Completo", "Código Cargo", "Puntos"]
    ws.append(headers)

    # Estilo para encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos de prueba con diferentes casos
    datos = [
        # CASO 1: Coincidencias perfectas (estos deberían matchear sin problemas)
        ["12345", "García, Juan", "757", 18000],
        ["67890", "López, María", "423", 25000],
        ["11111", "Fernández, Carlos", "851", 6000],

        # CASO 2: Diferencia de puntos (existe en sistema con 50000, pero aquí aparece con 45000)
        ["22222", "Martínez, Ana", "342", 45000],

        # CASO 3: Pago sin autorización (docente no existe en sistema)
        ["99999", "Pérez, Roberto", "555", 20000],

        # CASO 4: Otro pago sin autorización (cargo diferente para docente existente)
        ["12345", "García, Juan", "888", 10000],

        # NOTA: El cargo 654 de Rodríguez, Luis NO está en este Excel
        # Esto generará: "No se pagó cargo autorizado"
    ]

    for fila in datos:
        ws.append(fila)

    # Ajustar anchos de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12

    # Alinear datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].alignment = Alignment(horizontal="center")  # Legajo
        row[1].alignment = Alignment(horizontal="left")     # Nombre
        row[2].alignment = Alignment(horizontal="center")   # Código
        row[3].alignment = Alignment(horizontal="right")    # Puntos

    # Guardar archivo
    filename = "liquidacion_noviembre_2025_prueba.xlsx"
    wb.save(filename)

    print(f"✓ Excel de prueba creado: {filename}")
    print(f"\n=== CONTENIDO DEL EXCEL ===")
    print(f"Total de registros: {len(datos)}")
    print(f"\nDetalles:")

    print("\n1. COINCIDENCIAS ESPERADAS (3):")
    print("   - García, Juan (12345) - Cargo 757 - 18000 pts")
    print("   - López, María (67890) - Cargo 423 - 25000 pts")
    print("   - Fernández, Carlos (11111) - Cargo 851 - 6000 pts")

    print("\n2. DIFERENCIA DE PUNTOS (1):")
    print("   - Martínez, Ana (22222) - Cargo 342")
    print("     Sistema: 50000 pts | Excel: 45000 pts | Diferencia: -5000 pts")

    print("\n3. PAGO SIN AUTORIZACIÓN (2):")
    print("   - Pérez, Roberto (99999) - Docente no existe en sistema")
    print("   - García, Juan (12345) - Cargo 888 no autorizado")

    print("\n4. NO SE PAGÓ CARGO AUTORIZADO (1):")
    print("   - Rodríguez, Luis (33333) - Cargo 654 - 12000 pts")
    print("     (Está LIQUIDADO en sistema pero NO aparece en Excel)")

    print(f"\n=== RESUMEN ESPERADO ===")
    print(f"Total registros procesados: {len(datos)}")
    print(f"Coincidencias esperadas: 3")
    print(f"Discrepancias esperadas: 4")
    print(f"  - Diferencia de puntos: 1")
    print(f"  - Pago sin autorización: 2")
    print(f"  - No se pagó autorizado: 1")

    return filename

if __name__ == '__main__':
    filename = crear_excel_prueba()
    print(f"\n✓ Archivo listo para usar en Django Admin!")
    print(f"✓ Ubicación: /home/josehlo/src/face/{filename}")
